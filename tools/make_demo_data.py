"""Generates the fully synthetic demo dataset for the public showcase.

Writes two files into demo/:

  Demo Operational Record.xlsx  -- an 11-sheet workbook (Oct 2025 - Aug
      2026) matching the legacy layout importer.py expects: month/year in
      D6/E6, a daily block whose day-1 row varies per sheet, snapshot
      rows 2-3 with free-text timestamps, the mains formula string in
      L51, per-month chemical config in N51/O53:R53/O54, legacy-note
      triples in rows 57-62, and service datetimes in columns O-X.

  expected_figures.json -- every period's balance figures computed by
      this script's own independent arithmetic (mirroring
      docs/DATA_SPEC.md section 2), so tests can assert the import ->
      balance pipeline reproduces them without referencing any real data.

Everything here is invented. No real reading, comment, tank level, or
config value from any actual site appears; the story (a flowmeter
drifting past the watch threshold with one offline month, back-estimated
PLC-freeze days, and a mis-keyed mains snapshot) exists so every feature
of the tool has something to demonstrate.

Deterministic: same output on every run (seeded RNG, fixed story).
"""

import calendar
import json
import random
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl

OUT_DIR = Path("demo")
WORKBOOK_PATH = OUT_DIR / "Demo Operational Record.xlsx"
FIGURES_PATH = OUT_DIR / "expected_figures.json"

rng = random.Random(1)

# --- chemical/plant config eras (all values invented) -----------------------

CONFIG_A = dict(batch_volume_kl=5.10, base=3.20, acid=1.00, ferrous=0.92,
                h2o2=1.24, caustic=1.05)
# From April 2026 the demo plant re-estimated its dosing -- exercises
# config_history versioning exactly like a real mid-series change.
CONFIG_B = dict(batch_volume_kl=5.25, base=3.10, acid=1.00, ferrous=0.95,
                h2o2=1.18, caustic=1.02)

SERVICE_ITEMS = [
    "AE pump grease", "Air compressor drain", "Bag filter swap",
    "Dosing pump check", "pH probe calibration", "Flow meter verify",
    "Sight glass clean", "Cabinet filter", "Oil level check", "Alarm lamp test",
]
SERVICE_COLS = ["O", "P", "Q", "R", "S", "T", "U", "V", "W", "X"]

MAINS_START = 1650.0
START_TANKS = {"t02": 2.40, "t31": 4.10, "t32": 3.20, "t5": 1.80}

# Entered post-import by tools/setup_demo.py: a mis-keyed mains reading
# (1800.4 typed for 1700.4) left unresolved, so `gtp check` has a live
# V9 BLOCK to demonstrate the override workflow on.
TYPO_SNAPSHOT = {
    "taken_at": "2026-08-16T10:00:00",
    "mains_meter": 1800.4,
    "note": "site visit meter read",
}


@dataclass
class MonthSpec:
    year: int
    month: int
    sheet: str
    day1_row: int
    target_pct: float | None  # under-reading % of total_out; None = offline month
    mains_used: float
    config: dict
    boundary_text: str        # end-of-month snapshot free text (shared with next sheet)
    days_filled: int | None = None  # None = whole month (Aug is partial)
    other_adjustments: float = 0.0
    other_note: str | None = None


MONTHS = [
    MonthSpec(2025, 10, "Oct 25", 9, 0.12, 5.1, CONFIG_A, "31 Oct @ midnight"),
    MonthSpec(2025, 11, "Nov 25", 12, 0.16, 4.8, CONFIG_A, "30 Nov @ 1145pm"),
    MonthSpec(2025, 12, "Dec 25", 8, 0.14, 5.4, CONFIG_A, "31 Dec @ midnight"),
    MonthSpec(2026, 1, "Jan 26", 13, 0.21, 5.0, CONFIG_A, "31 Jan @ 0800am"),
    MonthSpec(2026, 2, "Feb 26", 10, None, 0.0, CONFIG_A, "28 Feb @ midnight"),
    MonthSpec(2026, 3, "Mar 26", 9, 0.18, 5.6, CONFIG_A, "31 Mar @ midnght"),  # sheet's own typo
    MonthSpec(2026, 4, "Apr 26", 14, 0.24, 4.9, CONFIG_B, "30 Apr @ 5pm"),
    MonthSpec(2026, 5, "May 26", 11, 0.30, 5.2, CONFIG_B, "31 May @ midnight"),
    MonthSpec(2026, 6, "Jun 26", 8, 0.38, 4.7, CONFIG_B, "30 Jun @ 0915am"),
    MonthSpec(2026, 7, "Jul 26", 12, 0.52, 5.5, CONFIG_B, "31 Jul @ midnight"),
    MonthSpec(2026, 8, "Aug 26", 10, 0.93, 2.1, CONFIG_B, "14 Aug @ 0930am",
              days_filled=14, other_adjustments=2.5,
              other_note="tanker top-up delivery"),
]

FIRST_BOUNDARY_TEXT = "30 Sep @ midnight"


@dataclass
class Day:
    date: str
    day: int
    reading: float | None
    error: float
    fit0501: float | None
    op_fraction: float | None
    comment: str | None = None
    bore_status: str | None = None
    estimated: bool = False  # derived by the importer from the PLC-freeze comment


@dataclass
class MonthData:
    spec: MonthSpec
    days: list[Day] = field(default_factory=list)
    total_in: float = 0.0
    total_out: float = 0.0
    start_tanks: dict = field(default_factory=dict)
    end_tanks: dict = field(default_factory=dict)
    mains_start: float = 0.0
    mains_end: float = 0.0
    chemicals_used: float = 0.0
    under_kl: float | None = None
    under_pct: float | None = None
    service: list[tuple[int, str, object]] = field(default_factory=list)  # (day, col, value)
    legacy_notes: list[tuple[str, float, float]] = field(default_factory=list)


def _make_days(spec: MonthSpec) -> list[Day]:
    """One Day per calendar day. The story overrides sit on fixed dates so
    the output is stable and the tests can name them.
    """
    n_in_month = calendar.monthrange(spec.year, spec.month)[1]
    filled = spec.days_filled if spec.days_filled is not None else n_in_month
    days: list[Day] = []
    for d in range(1, n_in_month + 1):
        date = f"{spec.year:04d}-{spec.month:02d}-{d:02d}"
        if d > filled:
            days.append(Day(date, d, None, 0.0, None, None))  # not yet entered
            continue
        if spec.target_pct is None:
            comment = None
            if d == 1:
                comment = "Plant OFFLINE - discharge line repair"
            elif d == n_in_month:
                comment = "Recommissioning checks pm"
            days.append(Day(date, d, 0.0, 0.0, 0.0, 0.0, comment=comment,
                            bore_status="All OFF"))
            continue

        reading = round(rng.uniform(33.0, 46.0), 1)
        fit0501 = round(max(reading + rng.uniform(-1.2, 2.0), 0.0), 1)
        day = Day(date, d, reading, 0.0, fit0501, 1.0, bore_status="All ON")
        days.append(day)

    # --- story overrides, by month ---------------------------------------
    key = (spec.year, spec.month)
    if key == (2025, 11):
        days[20].error = 0.4
        days[20].comment = "False flow during filter swap - 0.4 deducted"
    if key == (2026, 1):
        for d in range(11, 18):  # EW04 offline for a week -> lower inflow
            days[d].bore_status = "1 off - EW04 (screen clean)"
            days[d].reading = round(rng.uniform(29.0, 31.5), 1)
            days[d].fit0501 = round(days[d].reading + rng.uniform(-0.5, 1.0), 1)
        days[14].reading, days[14].fit0501 = 26.4, 27.1  # deliberate V4 warning
        days[14].comment = "Low day - EW04 still isolated"
    if key == (2026, 3):
        days[19].reading, days[19].fit0501 = 49.2, 47.8  # deliberate V3 warning
        days[19].comment = "Storm inflow overnight"
    if key == (2026, 4):
        days[21].error = 0.3
        days[21].comment = "False flow during filter swap - 0.3 deducted"
    if key == (2026, 5):
        # SCADA froze on the 11th; the 12th-13th repeat its values and are
        # back-estimated. The comment wording matters: the importer flags
        # is_estimated from a "PLC freeze" match.
        for d in (10, 11, 12):
            days[d].reading, days[d].fit0501 = 38.6, 39.0
        for d in (11, 12):
            days[d].comment = "PLC freeze - values back-filled from SCADA daily totals"
    if key == (2026, 8):
        days[8].comment = "PLC freeze overnight - day total back-filled from SCADA report"
        days[5].comment = "Tanker top-up delivery 2.5 kL (see adjustments)"

    # sprinkle routine comments on otherwise-quiet days
    if spec.target_pct is not None:
        quiet = [d for d in days if d.comment is None and d.reading is not None]
        for day in rng.sample(quiet, min(3, len(quiet))):
            day.comment = rng.choice([
                "s.v - routine site check", "Greased AE pump, all normal",
                "Heavy rain overnight, no issues", "Sample run collected",
            ])
    return days


def _split_tanks(total: float) -> dict:
    """Distribute a total volume across the four tanks, 2 dp, plausible."""
    t02 = round(total * rng.uniform(0.18, 0.26), 2)
    t31 = round(total * rng.uniform(0.30, 0.38), 2)
    t32 = round(total * rng.uniform(0.22, 0.30), 2)
    t5 = round(total - t02 - t31 - t32, 2)
    return {"t02": t02, "t31": t31, "t32": t32, "t5": t5}


def build_months() -> list[MonthData]:
    rng.seed(1)  # reset so repeated calls (e.g. from tests) are identical
    months: list[MonthData] = []
    tanks = dict(START_TANKS)
    mains = MAINS_START

    for spec in MONTHS:
        m = MonthData(spec=spec, days=_make_days(spec))
        filled = [d for d in m.days if d.reading is not None]
        # Same summation order and expressions as balance.py, so the
        # figures recorded here are exactly what the pipeline reproduces.
        m.total_in = sum(d.reading - d.error for d in filled)
        m.total_out = sum(d.fit0501 for d in filled)

        m.start_tanks = dict(tanks)
        m.mains_start = mains
        m.mains_end = round(mains + spec.mains_used, 1)

        cfg = spec.config
        clpb = cfg["base"] * (cfg["acid"] + cfg["ferrous"] + cfg["h2o2"] + cfg["caustic"])
        m.chemicals_used = (m.total_out / cfg["batch_volume_kl"]) * clpb / 1000

        mains_used_actual = m.mains_end - m.mains_start
        if spec.target_pct is None:
            m.end_tanks = dict(tanks)  # offline: nothing moved
        else:
            target_under = spec.target_pct / 100 * m.total_out
            start_volume = tanks["t02"] + tanks["t31"] + tanks["t32"] + tanks["t5"]
            end_volume_target = (
                start_volume + (m.total_in - m.total_out)
                + target_under + mains_used_actual + m.chemicals_used
                + spec.other_adjustments
            )
            m.end_tanks = _split_tanks(end_volume_target)

        # Recompute the achieved figures from the rounded stored values --
        # these, not the targets, go into expected_figures.json.
        start_volume = m.start_tanks["t02"] + m.start_tanks["t31"] + m.start_tanks["t32"] + m.start_tanks["t5"]
        end_volume = m.end_tanks["t02"] + m.end_tanks["t31"] + m.end_tanks["t32"] + m.end_tanks["t5"]
        discrepancy = end_volume - (start_volume + (m.total_in - m.total_out))
        m.under_kl = (discrepancy - mains_used_actual - m.chemicals_used
                      - spec.other_adjustments)
        m.under_pct = None if m.total_out == 0 else m.under_kl / m.total_out

        # service events: 2-3 per month, never on the day that lands on
        # row 15 (the importer skips that row as a legacy summary row)
        if spec.target_pct is not None:
            row15_day = 15 - spec.day1_row + 1
            candidates = [d.day for d in filled if d.day != row15_day]
            for day in rng.sample(candidates, min(3, len(candidates))):
                col = rng.choice(SERVICE_COLS)
                m.service.append((day, col, datetime(spec.year, spec.month, day, 9, 0)))
        else:
            m.service.append((10, "O", "OFFLINE"))  # exercises the unparsed-cell warning

        if spec.sheet in ("Oct 25", "Nov 25"):
            m.legacy_notes.append((
                "Hand balance (legacy)", round(m.under_kl, 1),
                round((m.under_pct or 0) * 100, 2),
            ))

        tanks = dict(m.end_tanks)
        mains = m.mains_end
        months.append(m)
    return months


MONTH_NAMES = {1: "January", 2: "February", 3: "March", 4: "April", 5: "May",
               6: "June", 7: "July", 8: "August", 9: "September", 10: "October",
               11: "November", 12: "December"}


def write_workbook(months: list[MonthData], path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.properties.creator = "GTP demo data generator"
    wb.properties.lastModifiedBy = "GTP demo data generator"

    prev_boundary_text = FIRST_BOUNDARY_TEXT
    for m in months:
        spec = m.spec
        ws = wb.create_sheet(spec.sheet)

        ws["A6"] = "Month:"
        ws["D6"] = MONTH_NAMES[spec.month]
        ws["E6"] = spec.year

        # snapshot rows: 2 = period start, 3 = period end. E..I are
        # t02, t31, t32, f4s, t5 -- f4s stays 0 in the demo.
        for row, tanks_row, text in (
            (2, m.start_tanks, prev_boundary_text),
            (3, m.end_tanks, spec.boundary_text),
        ):
            ws[f"D{row}"] = "Snapshot"
            ws[f"E{row}"] = tanks_row["t02"]
            ws[f"F{row}"] = tanks_row["t31"]
            ws[f"G{row}"] = tanks_row["t32"]
            ws[f"H{row}"] = 0.0
            ws[f"I{row}"] = tanks_row["t5"]
            ws[f"L{row}"] = text

        # column headers (cosmetic except O7:X7, which name service items)
        header_row = 7
        ws[f"B{header_row}"] = "Day"
        ws[f"C{header_row}"] = "Op"
        ws[f"D{header_row}"] = "Comments"
        ws[f"F{header_row}"] = "Error kL"
        ws[f"G{header_row}"] = "FIT0101 kL"
        ws[f"I{header_row}"] = "FIT0501 kL"
        ws[f"K{header_row}"] = "Bores"
        for col, item in zip(SERVICE_COLS, SERVICE_ITEMS):
            ws[f"{col}{header_row}"] = item

        for i, day in enumerate(m.days):
            row = spec.day1_row + i
            ws[f"B{row}"] = day.day
            if day.reading is None:
                continue  # future day: number only, no data yet
            ws[f"C{row}"] = day.op_fraction
            if day.comment:
                ws[f"D{row}"] = day.comment
            if day.error:
                ws[f"F{row}"] = day.error
            ws[f"G{row}"] = day.reading
            ws[f"I{row}"] = day.fit0501
            if day.bore_status:
                ws[f"K{row}"] = day.bore_status

        for day, col, value in m.service:
            ws[f"{col}{spec.day1_row + day - 1}"] = value

        if spec.other_adjustments:
            ws["L50"] = spec.other_note
            ws["M50"] = spec.other_adjustments
        # mains formula string, matching the legacy sheets' habit of a
        # trailing zero pro-rata term
        ws["L51"] = f"={m.mains_end}-{m.mains_start}+(0/30*5.3)"
        ws["N51"] = spec.config["batch_volume_kl"]
        ws["O53"] = spec.config["acid"]
        ws["P53"] = spec.config["ferrous"]
        ws["Q53"] = spec.config["h2o2"]
        ws["R53"] = spec.config["caustic"]
        ws["O54"] = spec.config["base"]

        for j, (label, kl, pct) in enumerate(m.legacy_notes):
            ws[f"G{57 + j}"] = label
            ws[f"L{57 + j}"] = kl
            ws[f"N{57 + j}"] = pct

        prev_boundary_text = spec.boundary_text

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def expected_figures(months: list[MonthData]) -> dict:
    periods = {}
    for m in months:
        spec = m.spec
        last = spec.days_filled or calendar.monthrange(spec.year, spec.month)[1]
        start = f"{spec.year:04d}-{spec.month:02d}-01"
        end = f"{spec.year:04d}-{spec.month:02d}-{last:02d}"
        periods[f"{start}..{end}"] = {
            "period_start": start,
            "period_end": end,
            "total_in": m.total_in,
            "total_out": m.total_out,
            "mains_used": m.mains_end - m.mains_start,
            "chemicals_used": m.chemicals_used,
            "other_adjustments": spec.other_adjustments,
            "under_kl": m.under_kl,
            "under_pct": m.under_pct,
            "contains_estimates": any(
                d.comment and "PLC freeze" in d.comment for d in m.days
            ),
        }
    return {
        "periods": periods,
        "sheets": len(months),
        "days_total": sum(len(m.days) for m in months),
        "distinct_snapshots": len(months) + 1,
        "estimated_dates": [
            d.date for m in months for d in m.days
            if d.comment and "PLC freeze" in d.comment
        ],
        "config_eras": {"2025-10-01": CONFIG_A, "2026-04-01": CONFIG_B},
        "typo_snapshot": dict(TYPO_SNAPSHOT),
    }


def check_story(months: list[MonthData]) -> None:
    """Assert the invented story actually plays out in the achieved
    figures (rounding shifts them slightly from the targets)."""
    pcts = {(m.spec.year, m.spec.month): m.under_pct for m in months}
    as_pct = lambda key: pcts[key] * 100

    assert pcts[(2026, 2)] is None, "offline month must have no percentage"
    assert as_pct((2026, 7)) > 0.5, "July must cross the watch threshold"
    assert 0.5 < as_pct((2026, 8)) < 1.0, "August must be WATCH, not ACTION"
    prior = [as_pct((2026, 5)), as_pct((2026, 6)), as_pct((2026, 7))]
    mean = sum(prior) / len(prior)
    assert as_pct((2026, 8)) > 2.0 * mean, "August must trigger ACCELERATING"
    for key in [(2025, 10), (2025, 11), (2025, 12), (2026, 1), (2026, 3),
                (2026, 4), (2026, 5), (2026, 6)]:
        assert as_pct(key) < 0.5, f"{key} must stay below the watch threshold"


def main() -> None:
    months = build_months()
    check_story(months)
    write_workbook(months, WORKBOOK_PATH)
    figures = expected_figures(months)
    FIGURES_PATH.write_text(json.dumps(figures, indent=2), encoding="utf-8")
    print(f"Wrote {WORKBOOK_PATH} ({len(months)} sheets)")
    print(f"Wrote {FIGURES_PATH}")
    for key, p in figures["periods"].items():
        pct = "n/a (offline)" if p["under_pct"] is None else f"{p['under_pct'] * 100:.3f}%"
        print(f"  {key}: under {pct}")


if __name__ == "__main__":
    main()
