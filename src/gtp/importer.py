"""Reads the legacy workbook and writes rows per docs/MIGRATION.md.

The daily data block does not start on the same row on every sheet, and
some figures (mains meter, other adjustments) are stored as formula
strings rather than plain numbers. Both are handled explicitly below —
see MIGRATION.md for the full column map and the reasoning.
"""

import ast
import calendar
import operator
import re
import sqlite3
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from gtp.db import DEFAULT_DB_PATH, init_db
from gtp.models import BalanceSnapshot, DailyReading, LegacyBalanceNote, ServiceEvent

DATE_COL = "B"
OP_COL = "C"
COMMENT_COL = "D"
ERROR_COL = "F"
READING_COL = "G"
OUT_COL = "I"
BORE_COL = "K"

# The 10 service/visit item columns, per MIGRATION.md. Columns M ("VISITS")
# and N ("Who") are excluded — they hold general visit info, not one of the
# named checklist items, and MIGRATION.md's column list omits them too.
SERVICE_COLUMNS = ["O", "P", "Q", "R", "S", "T", "U", "V", "W", "X"]

# Row 15 always holds a derived "Last serviced" summary, regardless of
# where a given sheet's day-1 row falls. Confirmed against the real file:
# every sheet has literal text "Last" in N15. Not real per-day data.
SERVICE_SUMMARY_ROW = 15

LEGACY_NOTE_ROWS = range(57, 63)  # rows 57-62: label/kL/pct triples

PLC_FREEZE_RE = re.compile(r"plc\s*freeze", re.IGNORECASE)

_SAFE_OPS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.UAdd: operator.pos,
    ast.USub: operator.neg,
}


def safe_eval_arithmetic(expr: str) -> float:
    """Evaluate a simple `+ - * / ()` arithmetic expression safely.

    Used for the workbook's formula-string cells (e.g. the mains meter
    formula), which must be read as text and evaluated ourselves rather
    than trusted from Excel's cached result.
    """
    text = expr.lstrip("=").strip()
    try:
        node = ast.parse(text, mode="eval")
    except SyntaxError as e:
        raise ValueError(f"not a valid arithmetic expression: {expr!r}") from e

    def _eval(n: ast.AST) -> float:
        if isinstance(n, ast.Expression):
            return _eval(n.body)
        if isinstance(n, ast.Constant) and isinstance(n.value, (int, float)):
            return n.value
        if isinstance(n, ast.BinOp) and type(n.op) in _SAFE_OPS:
            return _SAFE_OPS[type(n.op)](_eval(n.left), _eval(n.right))
        if isinstance(n, ast.UnaryOp) and type(n.op) in _SAFE_OPS:
            return _SAFE_OPS[type(n.op)](_eval(n.operand))
        raise ValueError(f"unsupported expression: {expr!r}")

    return _eval(node)


def find_day1_row(ws: Worksheet) -> int:
    """Scan column B from row 6 downward for the first day-1 cell.

    Per MIGRATION.md: do not hardcode row offsets, the daily block starts
    on a different row on every sheet.
    """
    for r in range(6, 60):
        val = ws[f"{DATE_COL}{r}"].value
        if isinstance(val, (int, float)) and float(val).is_integer() and 1 <= val <= 31:
            return r
    raise ValueError(f"could not find a day-1 row on sheet {ws.title!r}")


_TIME_RE = re.compile(r"(\d{1,4})\s*(am|pm)", re.IGNORECASE)
_DATE_RE = re.compile(r"(\d{1,2})\s+([A-Za-z]+)")


def _parse_time_of_day(text: str) -> tuple[str, bool]:
    lowered = text.lower()
    if "midn" in lowered:  # catches "midnight" and the sheet's own "midnght" typo
        return "00:00:00", True
    m = _TIME_RE.search(lowered)
    if m:
        digits, ampm = m.group(1), m.group(2)
        if len(digits) <= 2:
            hour, minute = int(digits), 0
        else:
            hour, minute = int(digits[:-2]), int(digits[-2:])
        if ampm == "pm" and hour != 12:
            hour += 12
        if ampm == "am" and hour == 12:
            hour = 0
        return f"{hour:02d}:{minute:02d}:00", True
    return "00:00:00", False


def parse_snapshot_time(text: str, month: int, year: int) -> tuple[str, str | None]:
    """Parse a free-text snapshot timestamp like "30 Jun @ midnight".

    Returns (iso_datetime, note). note is None on a clean parse; otherwise
    it explains what couldn't be parsed, and a midnight default is used —
    per MIGRATION.md's "parse best-effort; on failure, store the raw
    string and default the time to midnight with a flag."
    """
    m = _DATE_RE.search(text)
    if not m:
        # Total parse failure: fall back to day 1 of the sheet's own month.
        return f"{year:04d}-{month:02d}-01T00:00:00", f"date parse failed, raw: {text!r}"

    day = int(m.group(1))
    month_name = m.group(2)[:3].title()
    try:
        parsed_month = datetime.strptime(month_name, "%b").month
    except ValueError:
        return f"{year:04d}-{month:02d}-01T00:00:00", f"month parse failed, raw: {text!r}"

    if parsed_month == month:
        year_out = year
    elif parsed_month == month - 1 or (month == 1 and parsed_month == 12):
        year_out = year - 1 if (month == 1 and parsed_month == 12) else year
    elif parsed_month == month + 1 or (month == 12 and parsed_month == 1):
        year_out = year + 1 if (month == 12 and parsed_month == 1) else year
    else:
        year_out = year  # unexpected spillover; best-effort fallback

    time_str, time_ok = _parse_time_of_day(text)
    iso = f"{year_out:04d}-{parsed_month:02d}-{day:02d}T{time_str}"
    note = None if time_ok else f"time parse failed, raw: {text!r}"
    return iso, note


_MAINS_RE = re.compile(r"^=?\s*([\d.]+)\s*-\s*([\d.]+)")


def _parse_mains_readings(formula: str) -> tuple[float, float]:
    """Pull the two literal meter readings out of the mains formula string.

    Returns (end_reading, start_reading).
    """
    m = _MAINS_RE.match(formula.strip())
    if not m:
        raise ValueError(f"could not parse mains formula: {formula!r}")
    return float(m.group(1)), float(m.group(2))


@dataclass
class SheetImportResult:
    sheet_name: str
    year: int
    month: int
    daily_readings: list[DailyReading]
    start_snapshot: BalanceSnapshot
    end_snapshot: BalanceSnapshot
    service_events: list[ServiceEvent]
    legacy_notes: list[LegacyBalanceNote]
    batch_volume_kl: float
    chem_base_litres_per_batch: float
    chem_acid_factor: float
    chem_ferrous_factor: float
    chem_h2o2_factor: float
    chem_caustic_factor: float
    mains_used_reported: float
    warnings: list[str] = field(default_factory=list)


def _read_numeric_cell(ws: Worksheet, ws_values: Worksheet, coord: str) -> float | None:
    """Read a cell that should be a plain number but is occasionally a
    formula. Literal arithmetic (e.g. "=1690.5-1684.4") is evaluated
    directly; formulas that reference other cells (e.g. "=G13*E$11",
    found in early-January's error column, which safe_eval_arithmetic
    can't resolve) fall back to Excel's own cached result instead.
    """
    raw = ws[coord].value
    if raw is None:
        return None
    if isinstance(raw, str) and raw.startswith("="):
        try:
            return safe_eval_arithmetic(raw)
        except ValueError:
            cached = ws_values[coord].value
            return None if cached is None else float(cached)
    return float(raw)


def parse_sheet(ws: Worksheet, ws_values: Worksheet) -> SheetImportResult:
    month_name = ws["D6"].value
    year = int(ws["E6"].value)
    month = datetime.strptime(str(month_name)[:3].title(), "%b").month

    day1_row = find_day1_row(ws)
    days_in_month = calendar.monthrange(year, month)[1]

    warnings: list[str] = []
    daily_readings: list[DailyReading] = []
    row = day1_row
    expected_day = 1
    while True:
        val = ws[f"{DATE_COL}{row}"].value
        if not (isinstance(val, (int, float)) and float(val).is_integer()):
            break
        day = int(val)
        if day != expected_day:
            raise ValueError(
                f"sheet {ws.title!r}: expected day {expected_day} at row {row}, got {day}"
            )
        date = f"{year:04d}-{month:02d}-{day:02d}"
        comment = ws[f"{COMMENT_COL}{row}"].value
        error = _read_numeric_cell(ws, ws_values, f"{ERROR_COL}{row}") or 0.0
        reading = _read_numeric_cell(ws, ws_values, f"{READING_COL}{row}")
        fit0501 = _read_numeric_cell(ws, ws_values, f"{OUT_COL}{row}")
        op_fraction = ws[f"{OP_COL}{row}"].value
        bore_status = ws[f"{BORE_COL}{row}"].value

        is_estimated = 0
        estimate_reason = None
        if comment and PLC_FREEZE_RE.search(str(comment)):
            is_estimated = 1
            estimate_reason = f"PLC freeze noted in source comment: {comment}"

        daily_readings.append(
            DailyReading(
                date=date,
                fit0101_reading=reading,
                fit0101_error=error,
                fit0501=fit0501,
                op_fraction=op_fraction,
                bore_status=bore_status,
                comment=comment,
                is_estimated=is_estimated,
                estimate_reason=estimate_reason,
                entered_at=datetime.now().isoformat(timespec="seconds"),
            )
        )
        row += 1
        expected_day += 1

    n_days = len(daily_readings)
    if n_days != days_in_month:
        raise ValueError(
            f"sheet {ws.title!r}: found {n_days} day rows, "
            f"expected {days_in_month} for {year}-{month:02d}"
        )

    start_snap = _parse_snapshot(ws, row_num=2, month=month, year=year, warnings=warnings)
    end_snap = _parse_snapshot(ws, row_num=3, month=month, year=year, warnings=warnings)

    mains_formula = ws["L51"].value
    end_reading, start_reading = _parse_mains_readings(mains_formula)
    start_snap.mains_meter = start_reading
    end_snap.mains_meter = end_reading
    # The true mains_used for THIS sheet's own period, evaluated directly
    # from its own formula rather than derived from the shared start/end
    # snapshot rows later. Needed because adjacent sheets occasionally
    # disagree by ~0.1 kL about the reading at a shared boundary moment
    # (confirmed in the real data at the May/Jun and Jul/Aug boundaries) —
    # deriving mains_used from the merged snapshot table would then use
    # whichever sheet was imported last, which isn't necessarily the one
    # this month's own documented balance figure was computed from.
    mains_used_reported = safe_eval_arithmetic(mains_formula)

    other_total = 0.0
    for col in "MNOPQ":
        cell = ws[f"{col}50"].value
        if cell is None:
            continue
        other_total += (
            safe_eval_arithmetic(str(cell)) if isinstance(cell, str) else float(cell)
        )
    end_snap.other_adjustments = other_total

    batch_volume_kl = float(ws["N51"].value)
    chem_acid, chem_ferrous, chem_h2o2, chem_caustic = (
        float(ws[f"{c}53"].value) for c in "OPQR"
    )
    chem_base = float(ws["O54"].value)

    legacy_notes = _parse_legacy_notes(ws)
    service_events, service_warnings = _parse_service_events(ws, day1_row, n_days)
    warnings.extend(service_warnings)

    return SheetImportResult(
        sheet_name=ws.title,
        year=year,
        month=month,
        daily_readings=daily_readings,
        start_snapshot=start_snap,
        end_snapshot=end_snap,
        service_events=service_events,
        legacy_notes=legacy_notes,
        batch_volume_kl=batch_volume_kl,
        chem_base_litres_per_batch=chem_base,
        chem_acid_factor=chem_acid,
        chem_ferrous_factor=chem_ferrous,
        chem_h2o2_factor=chem_h2o2,
        chem_caustic_factor=chem_caustic,
        mains_used_reported=mains_used_reported,
        warnings=warnings,
    )


def _as_float(value) -> float | None:
    """Tank-level cells are normally plain numbers but are occasionally
    formula strings (e.g. F4s "=5.115+2.23") — evaluate rather than store
    the formula text.
    """
    if value is None:
        return None
    return safe_eval_arithmetic(str(value)) if isinstance(value, str) else float(value)


def _parse_snapshot(
    ws: Worksheet, row_num: int, month: int, year: int, warnings: list[str]
) -> BalanceSnapshot:
    t02, t31, t32, f4s, t5 = (
        _as_float(ws[f"{c}{row_num}"].value) for c in "EFGHI"
    )
    time_text = ws[f"L{row_num}"].value or ""
    taken_at, note = parse_snapshot_time(str(time_text), month, year)
    if note:
        warnings.append(f"{ws.title}: {note}")
    return BalanceSnapshot(
        taken_at=taken_at,
        t02=t02,
        t31=t31,
        t32=t32,
        f4s=f4s or 0.0,
        t5=t5,
        note=note,
    )


def _parse_legacy_notes(ws: Worksheet) -> list[LegacyBalanceNote]:
    notes = []
    for row in LEGACY_NOTE_ROWS:
        label = ws[f"G{row}"].value
        kl = ws[f"L{row}"].value
        pct = ws[f"N{row}"].value
        if label is None and kl is None and pct is None:
            continue
        notes.append(
            LegacyBalanceNote(sheet=ws.title, row=row, label=label, kl=kl, pct=pct)
        )
    return notes


def _parse_service_events(
    ws: Worksheet, day1_row: int, n_days: int
) -> tuple[list[ServiceEvent], list[str]]:
    headers = {col: ws[f"{col}7"].value for col in SERVICE_COLUMNS}
    events: list[ServiceEvent] = []
    warnings: list[str] = []
    for row in range(day1_row, day1_row + n_days):
        if row == SERVICE_SUMMARY_ROW:
            continue
        for col in SERVICE_COLUMNS:
            val = ws[f"{col}{row}"].value
            if val is None:
                continue
            if isinstance(val, datetime):
                events.append(
                    ServiceEvent(date=val.date().isoformat(), item=headers[col])
                )
            else:
                warnings.append(
                    f"{ws.title} {col}{row}: unparsed service cell {val!r} "
                    f"(item {headers[col]!r})"
                )
    return events, warnings


def parse_workbook(path: str | Path) -> list[SheetImportResult]:
    """Parse every sheet in the legacy workbook. Read-only, no database access."""
    wb = openpyxl.load_workbook(path, data_only=False)
    wb_values = openpyxl.load_workbook(path, data_only=True)
    return [parse_sheet(wb[name], wb_values[name]) for name in wb.sheetnames]


@dataclass
class ImportSummary:
    sheets_imported: int
    days_imported: int
    snapshots_imported: int
    service_events_imported: int
    legacy_notes_imported: int
    config_rows_imported: int
    warnings: list[str]


def import_workbook(
    path: str | Path, db_path: str | Path = DEFAULT_DB_PATH
) -> ImportSummary:
    """Parse the workbook and write it into the database. Safe to re-run."""
    sheets = parse_workbook(path)
    conn = init_db(db_path)

    warnings: list[str] = []
    snapshots_written = 0
    config_rows_written = 0

    # service_event and legacy_balance_note are entirely derived from this
    # read-only workbook, so a full replace on each run is simplest and
    # keeps re-imports idempotent without a merge step.
    conn.execute("DELETE FROM service_event")
    conn.execute("DELETE FROM legacy_balance_note")

    for sheet in sheets:
        warnings.extend(sheet.warnings)

        for r in sheet.daily_readings:
            conn.execute(
                """
                INSERT INTO daily_reading
                    (date, fit0101_reading, fit0101_error, fit0501, op_fraction,
                     bore_status, comment, is_estimated, estimate_reason, entered_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(date) DO UPDATE SET
                    fit0101_reading = excluded.fit0101_reading,
                    fit0101_error   = excluded.fit0101_error,
                    fit0501         = excluded.fit0501,
                    op_fraction     = excluded.op_fraction,
                    bore_status     = excluded.bore_status,
                    comment         = excluded.comment,
                    is_estimated    = excluded.is_estimated,
                    estimate_reason = excluded.estimate_reason,
                    entered_at      = excluded.entered_at
                """,
                (
                    r.date, r.fit0101_reading, r.fit0101_error, r.fit0501,
                    r.op_fraction, r.bore_status, r.comment, r.is_estimated,
                    r.estimate_reason, r.entered_at,
                ),
            )

        for snap in (sheet.start_snapshot, sheet.end_snapshot):
            conn.execute(
                """
                INSERT INTO balance_snapshot
                    (taken_at, t02, t31, t32, t5, f4s, mains_meter,
                     other_adjustments, other_note, note)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(taken_at) DO UPDATE SET
                    t02               = excluded.t02,
                    t31               = excluded.t31,
                    t32               = excluded.t32,
                    t5                = excluded.t5,
                    f4s               = excluded.f4s,
                    mains_meter       = excluded.mains_meter,
                    other_adjustments = excluded.other_adjustments,
                    other_note        = excluded.other_note,
                    note              = excluded.note
                """,
                (
                    snap.taken_at, snap.t02, snap.t31, snap.t32, snap.t5, snap.f4s,
                    snap.mains_meter, snap.other_adjustments, snap.other_note, snap.note,
                ),
            )
            snapshots_written += 1

        for ev in sheet.service_events:
            conn.execute(
                "INSERT INTO service_event (date, item, note) VALUES (?, ?, ?)",
                (ev.date, ev.item, ev.note),
            )

        for note in sheet.legacy_notes:
            conn.execute(
                "INSERT INTO legacy_balance_note (sheet, row, label, kl, pct) "
                "VALUES (?, ?, ?, ?, ?)",
                (note.sheet, note.row, note.label, note.kl, note.pct),
            )

        effective_from = f"{sheet.year:04d}-{sheet.month:02d}-01"
        conn.execute(
            """
            INSERT INTO config_history
                (effective_from, batch_volume_kl, chem_base_litres_per_batch,
                 chem_acid_factor, chem_ferrous_factor, chem_h2o2_factor,
                 chem_caustic_factor, mains_used_reported, source)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(effective_from) DO UPDATE SET
                batch_volume_kl            = excluded.batch_volume_kl,
                chem_base_litres_per_batch = excluded.chem_base_litres_per_batch,
                chem_acid_factor           = excluded.chem_acid_factor,
                chem_ferrous_factor        = excluded.chem_ferrous_factor,
                chem_h2o2_factor           = excluded.chem_h2o2_factor,
                chem_caustic_factor        = excluded.chem_caustic_factor,
                mains_used_reported        = excluded.mains_used_reported,
                source                     = excluded.source
            """,
            (
                effective_from, sheet.batch_volume_kl, sheet.chem_base_litres_per_batch,
                sheet.chem_acid_factor, sheet.chem_ferrous_factor, sheet.chem_h2o2_factor,
                sheet.chem_caustic_factor, sheet.mains_used_reported, f"legacy:{sheet.sheet_name}",
            ),
        )
        config_rows_written += 1

    conn.commit()

    (total_events,) = conn.execute("SELECT COUNT(*) FROM service_event").fetchone()
    (total_notes,) = conn.execute("SELECT COUNT(*) FROM legacy_balance_note").fetchone()

    return ImportSummary(
        sheets_imported=len(sheets),
        days_imported=sum(len(s.daily_readings) for s in sheets),
        snapshots_imported=snapshots_written,
        service_events_imported=total_events,
        legacy_notes_imported=total_notes,
        config_rows_imported=config_rows_written,
        warnings=warnings,
    )
